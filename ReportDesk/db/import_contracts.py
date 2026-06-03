"""从 Excel 导入工程合同名录（项目名称、负责人、经办人）。"""

from __future__ import annotations

from pathlib import Path
from typing import Any

# 表头别名（精确匹配列名）
_COL_PROJECT = "项目名称"
_COL_MANAGER = "负责人"
_COL_HANDLER = "经办人"


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def read_contract_rows(excel_path: Path) -> list[dict[str, str]]:
    from openpyxl import load_workbook

    path = Path(excel_path)
    if not path.is_file():
        raise FileNotFoundError(f"Excel 不存在: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        wb.close()
        return []

    headers = [_cell_str(h) for h in header_row]
    col_map: dict[str, int] = {}
    for i, h in enumerate(headers):
        if h in (_COL_PROJECT, _COL_MANAGER, _COL_HANDLER):
            col_map[h] = i

    if _COL_PROJECT not in col_map:
        wb.close()
        raise ValueError(f"Excel 须包含列「{_COL_PROJECT}」，当前表头: {headers}")

    out: list[dict[str, str]] = []
    for row in rows_iter:
        if not row:
            continue
        cells = list(row)
        name = _cell_str(cells[col_map[_COL_PROJECT]] if col_map[_COL_PROJECT] < len(cells) else "")
        if not name:
            continue
        manager = ""
        handler = ""
        if _COL_MANAGER in col_map and col_map[_COL_MANAGER] < len(cells):
            manager = _cell_str(cells[col_map[_COL_MANAGER]])
        if _COL_HANDLER in col_map and col_map[_COL_HANDLER] < len(cells):
            handler = _cell_str(cells[col_map[_COL_HANDLER]])
        out.append(
            {
                "project_name": name,
                "manager": manager,
                "handler": handler,
            }
        )
    wb.close()
    return out
